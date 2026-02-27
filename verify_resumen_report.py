import requests
import openpyxl
import io
import os

def test_resumen_report():
    # Attempt to get an existing taxpayer NIT and year from DB
    nit = "6779991019"
    year = 2025
    
    url = f"http://localhost:8000/forms/reports/resumen-200/excel?nit={nit}&year={year}"
    print(f"Testing URL: {url}")
    
    try:
        response = requests.get(url)
        if response.status_code == 404:
            print("No data found for this NIT/Year. Trying to find any NIT in DB...")
            taxpayers = requests.get("http://localhost:8000/forms/taxpayers").json()
            if taxpayers:
                nit = taxpayers[0]['nit']
                url = f"http://localhost:8000/forms/reports/resumen-200/excel?nit={nit}&year={year}"
                print(f"Retrying with NIT: {nit}")
                response = requests.get(url)
        
        if response.status_code != 200:
            print(f"Error: Status code {response.status_code}")
            print(response.text)
            return
            
        print("Report generated successfully!")
        
        # Verify Excel content
        content = io.BytesIO(response.content)
        wb = openpyxl.load_workbook(content)
        ws = wb.active
        
        print(f"Active Sheet: {ws.title}")
        
        # Check some values
        # Row 3 should have 'C13'
        c13_val = ws.cell(row=3, column=2).value
        print(f"Row 3, Col 2 (Expected C13): {c13_val}")
        
        # Check summary labels
        labels = [ws.cell(row=3, column=j).value for j in range(52, 58)]
        print(f"Summary labels (Cols 52-57): {labels}")
        
        # Check if month rows have some data (non-zero or float)
        first_row_val = ws.cell(row=4, column=2).value
        print(f"Row 4, Col 2 (C13 value for Jan): {first_row_val}")
        
    except Exception as e:
        print(f"Verification failed: {e}")

if __name__ == "__main__":
    test_resumen_report()
