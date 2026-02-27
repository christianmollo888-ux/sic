from fpdf import FPDF
from datetime import date
import io
import re
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side

ILLEGAL_CHARACTERS_RE = re.compile(
    r'[\000-\010]|[\013-\014]|[\016-\037]|[\x00-\x1f\x7f-\x9f]|[\ufffe-\uffff]'
)

def clean_string(s):
    if not isinstance(s, str):
        return s
    return ILLEGAL_CHARACTERS_RE.sub("", s)

class BSSPDF(FPDF):
    def __init__(self, *args, **kwargs):
        self.end_date_str = kwargs.pop('end_date_str', '')
        super().__init__(*args, **kwargs)

    def header(self):
        # Company Info
        self.set_font('helvetica', 'B', 10)
        self.cell(0, 5, 'SISTEMA CONTABLE SIC4BUS', ln=True)
        self.set_font('helvetica', '', 8)
        self.cell(0, 4, 'Reporte Generado Automáticamente', ln=True)
        self.ln(5)
        
        # Title
        self.set_font('helvetica', 'B', 14)
        self.cell(0, 10, 'BALANCE DE SUMAS Y SALDOS', align='C', ln=True)
        self.set_font('helvetica', '', 10)
        self.cell(0, 5, f'Al {self.end_date_str}', align='C', ln=True)
        self.ln(10)
        
        # Table Header
        self.set_font('helvetica', 'B', 8)
        self.set_fill_color(230, 230, 230)
        self.cell(25, 7, 'CODIGO', 1, 0, 'C', True)
        self.cell(85, 7, 'CUENTA', 1, 0, 'C', True)
        self.cell(20, 7, 'DEBE', 1, 0, 'C', True)
        self.cell(20, 7, 'HABER', 1, 0, 'C', True)
        self.cell(20, 7, 'DEUDOR', 1, 0, 'C', True)
        self.cell(20, 7, 'ACREEDOR', 1, 1, 'C', True)

    def footer(self):
        self.set_y(-15)
        self.set_font('helvetica', 'I', 8)
        self.cell(0, 10, f'Página {self.page_no()}/{{nb}}', align='C')

def generate_bss_pdf(data, end_date_str):
    pdf = BSSPDF(end_date_str=end_date_str)
    pdf.alias_nb_pages()
    pdf.add_page()
    
    pdf.set_font('helvetica', '', 8)
    
    for row in data:
        # Highlight parent accounts (bold)
        if row['level'] is not None and row['level'] < 5:
            pdf.set_font('helvetica', 'B', 8)
        else:
            pdf.set_font('helvetica', '', 8)
            
        # Draw cells
        pdf.cell(25, 6, str(row['code']), 1)
        # Truncate name if too long
        name = row['name']
        if len(name) > 55:
            name = name[:52] + "..."
        pdf.cell(85, 6, name, 1)
        
        # Numbers formatting
        pdf.cell(20, 6, f"{row['debit']:,.2f}", 1, 0, 'R')
        pdf.cell(20, 6, f"{row['credit']:,.2f}", 1, 0, 'R')
        pdf.cell(20, 6, f"{row['deudor']:,.2f}", 1, 0, 'R')
        pdf.cell(20, 6, f"{row['acreedor']:,.2f}", 1, 1, 'R')
        
    # Calculate totals
    total_debit = sum(row['debit'] for row in data if row['level'] == 5)
    total_credit = sum(row['credit'] for row in data if row['level'] == 5)
    total_deudor = sum(row['deudor'] for row in data if row['level'] == 5)
    total_acreedor = sum(row['acreedor'] for row in data if row['level'] == 5)
    
    pdf.set_font('helvetica', 'B', 8)
    pdf.cell(110, 6, "TOTALES", 1, 0, 'C')
    pdf.cell(20, 6, f"{total_debit:,.2f}", 1, 0, 'R')
    pdf.cell(20, 6, f"{total_credit:,.2f}", 1, 0, 'R')
    pdf.cell(20, 6, f"{total_deudor:,.2f}", 1, 0, 'R')
    pdf.cell(20, 6, f"{total_acreedor:,.2f}", 1, 1, 'R')
        
    return pdf.output()

# ====== BALANCE GENERAL ======

class BalanceGeneralPDF(FPDF):
    def __init__(self, *args, **kwargs):
        self.start_date_str = kwargs.pop('start_date_str', '')
        self.end_date_str = kwargs.pop('end_date_str', '')
        super().__init__(*args, **kwargs)
        self.set_margins(15, 15, 15)

    def header(self):
        # Company name (left)
        self.set_font('helvetica', 'B', 10)
        self.cell(0, 5, 'COMSATELITAL S.R.L.', ln=True, align='L')

        # Title centered
        self.set_font('helvetica', 'B', 13)
        self.cell(0, 6, 'BALANCE GENERAL', align='C', ln=True)

        # Date range centered
        self.set_font('helvetica', '', 9)
        self.cell(0, 5, f'DEL  {self.start_date_str}  AL  {self.end_date_str}', align='C', ln=True)

        self.ln(2)
        # Column headers
        self.set_font('helvetica', 'B', 8)
        self.cell(35, 5, 'CODIGO', 0, 0, 'L')
        self.cell(110, 5, 'NOMBRE', 0, 0, 'L')
        self.cell(30, 5, 'BOLIVIANOS', 0, 1, 'R')
        self.set_font('helvetica', '', 7)
        self.cell(145, 4, '(Expresado en Bolivianos)', 0, 0, 'L')
        self.cell(30, 4, 'Bs.', 0, 1, 'R')

        # Separator line
        self.line(15, self.get_y(), 195, self.get_y())
        self.ln(2)

    def footer(self):
        self.set_y(-15)
        self.set_font('helvetica', 'I', 8)
        self.cell(0, 10, f'Página {self.page_no()}/{{nb}}', align='C')


def _get_bg_indent(level):
    indents = {1: '', 2: '  ', 3: '    ', 4: '      ', 5: '        '}
    return indents.get(level, '')


def generate_balance_general_pdf(data, start_date_str, end_date_str):
    pdf = BalanceGeneralPDF(start_date_str=start_date_str, end_date_str=end_date_str)
    pdf.alias_nb_pages()
    pdf.add_page()

    def fmt(val):
        v = float(val)
        if v < 0:
            return f'-{abs(v):,.2f}'
        return f'{v:,.2f}'

    def draw_row(code, name, balance, level):
        if pdf.get_y() > 265:
            pdf.add_page()

        if level is not None and level < 3:
            pdf.set_font('helvetica', 'B', 8)
        else:
            pdf.set_font('helvetica', '', 8)

        indent = _get_bg_indent(level)
        max_name = 62 - len(indent)
        name_trunc = name[:max_name-3] + '...' if len(name) > max_name else name
        display_name = indent + name_trunc

        pdf.cell(35, 5, str(code), 0, 0, 'L')
        pdf.cell(110, 5, display_name, 0, 0, 'L')
        bal_str = fmt(balance) if balance != 0 else ''
        pdf.cell(30, 5, bal_str, 0, 1, 'R')

    # ---- CUENTAS DE ACTIVO ----
    pdf.set_font('helvetica', 'B', 9)
    pdf.cell(0, 6, 'CUENTAS DE ACTIVO', 0, 1, 'L')

    for row in data['activo']:
        draw_row(row['code'], row['name'], row['balance'], row['level'])

    pdf.ln(2)
    pdf.set_font('helvetica', 'B', 9)
    pdf.line(15, pdf.get_y(), 195, pdf.get_y())
    pdf.ln(1)
    pdf.cell(145, 6, 'TOTAL CUENTAS DE ACTIVO', 0, 0, 'L')
    pdf.cell(30, 6, fmt(data['total_activo']), 0, 1, 'R')
    pdf.ln(5)

    # ---- CUENTAS DE PASIVO Y PATRIMONIO ----
    pdf.set_font('helvetica', 'B', 9)
    pdf.cell(0, 6, 'CUENTAS DE PASIVO Y PATRIMONIO', 0, 1, 'L')

    for row in data['pasivo_patrimonio']:
        draw_row(row['code'], row['name'], row['balance'], row['level'])

    pdf.ln(2)
    pdf.set_font('helvetica', 'B', 9)
    pdf.line(15, pdf.get_y(), 195, pdf.get_y())
    pdf.ln(1)
    pdf.cell(145, 6, 'TOTAL CUENTAS DE PASIVO Y PATRIMONIO', 0, 0, 'L')
    pdf.cell(30, 6, fmt(data['total_pasivo_patrimonio']), 0, 1, 'R')

    return pdf.output()

# ====== END BALANCE GENERAL ======

def get_voucher_type_name(entry_type_code):

    mapping = {
        '1': 'INGRESO',
        '2': 'EGRESO',
        '3': 'DIARIO',
        '11': 'AJUSTE',
    }
    return mapping.get(str(entry_type_code), 'CONTABLE')

class ComprobantePDF(FPDF):
    def __init__(self, *args, **kwargs):
        self.entry = kwargs.pop('entry', None)
        super().__init__(*args, **kwargs)
        self.set_margins(11, 15, 11)

    def header(self):
        if not self.entry:
            return
            
        self.rect(10, 10, 190, 277)
        self.rect(11, 11, 188, 275)
            
        self.set_y(15)
        self.set_x(13)
        self.set_font('helvetica', '', 9)
        self.cell(80, 4, 'COMSATELITAL S.R.L.', ln=True, align='L')
        self.set_x(13)
        self.cell(80, 4, 'LA PAZ-Bolivia', ln=True, align='L')
        
        self.set_xy(145, 15)
        self.set_font('helvetica', '', 8)
        self.cell(32, 4, 'Fecha', 1, 0, 'C')
        self.cell(15, 4, 'T.C.', 1, 1, 'C')
        
        self.set_x(145)
        d = self.entry.date
        self.cell(10.66, 4, str(d.day).zfill(2), 1, 0, 'C')
        self.cell(10.66, 4, str(d.month).zfill(2), 1, 0, 'C')
        self.cell(10.68, 4, str(d.year), 1, 0, 'C')
        self.cell(15, 4, '6.96000', 1, 1, 'C')
        
        self.set_xy(13, 30)
        type_name = get_voucher_type_name(self.entry.entry_type)
        self.set_font('helvetica', 'B', 14)
        self.cell(130, 10, f'COMPROBANTE DE  {type_name}', align='C')
        
        self.set_xy(155, 28)
        self.set_font('helvetica', 'B', 14)
        self.cell(38, 10, f'Nº {self.entry.entry_number or self.entry.id}', 1, 1, 'C', fill=False)
        
        self.set_y(42)
        self.line(11, 40, 199, 40)
        
        self.set_font('helvetica', 'B', 8)
        self.set_x(13)
        self.cell(30, 5, 'LUGAR Y FECHA :', 0, 0)
        self.set_font('helvetica', '', 8)
        months = ["Enero", "Febrero", "Marzo", "Abril", "Mayo", "Junio", "Julio", "Agosto", "Septiembre", "Octubre", "Noviembre", "Diciembre"]
        fecha_str = f"LA PAZ, {d.day} de {months[d.month-1]} de {d.year}"
        self.cell(0, 5, fecha_str, 0, 1)
        
        self.set_font('helvetica', 'B', 8)
        self.set_x(13)
        self.cell(30, 5, 'CONCEPTO      :', 0, 0)
        self.set_font('helvetica', '', 8)
        
        y_before_glosa = self.get_y()
        self.multi_cell(0, 5, str(self.entry.description or '').upper())
        y_after_glosa = self.get_y()
        self.set_y(max(y_after_glosa, y_before_glosa + 5))
        
        self.ln(2)
        self.line(11, self.get_y(), 199, self.get_y())
        self.line(11, self.get_y() + 1, 199, self.get_y() + 1)
        
        self.ln(1)
        y_th = self.get_y()
        self.set_font('helvetica', 'B', 8)
        
        self.cell(20, 5, '', 0, 0, 'C')
        self.cell(88, 5, 'NOMBRE CUENTA', 0, 0, 'C')
        self.cell(40, 5, 'BOLIVIANOS', 1, 0, 'C')
        self.cell(40, 5, 'DÓLARES', 1, 1, 'C')
        
        self.cell(20, 5, 'CODIGO', 0, 0, 'C')
        self.cell(88, 5, 'REFERENCIA', 0, 0, 'C')
        self.cell(20, 5, 'DEBE', 1, 0, 'C')
        self.cell(20, 5, 'HABER', 1, 0, 'C')
        self.cell(20, 5, 'DEBE', 1, 0, 'C')
        self.cell(20, 5, 'HABER', 1, 1, 'C')
        
        y_th_end = self.get_y()
        self.line(31, y_th, 31, y_th_end)
        self.line(119, y_th, 119, y_th_end)
        self.line(11, y_th_end, 199, y_th_end)

    def footer(self):
        self.set_y(-30)
        self.set_font('helvetica', '', 8)
        self.cell(63, 5, '_________________________', 0, 0, 'C')
        self.cell(63, 5, '_________________________', 0, 0, 'C')
        self.cell(64, 5, '_________________________', 0, 1, 'C')
        self.cell(63, 5, 'Preparado por', 0, 0, 'C')
        self.cell(63, 5, 'Revisado por', 0, 0, 'C')
        self.cell(64, 5, 'Autorizado por', 0, 1, 'C')
        
        self.set_y(-15)
        self.set_font('helvetica', 'I', 8)
        self.cell(0, 10, f'Página {self.page_no()}/{{nb}}', align='C')

def generate_comprobante_pdf(entry):
    pdf = ComprobantePDF(entry=entry)
    pdf.alias_nb_pages()
    pdf.add_page()
    
    pdf.set_font('helvetica', '', 8)
    
    total_debit = 0
    total_credit = 0
    total_debit_usd = 0
    total_credit_usd = 0
    
    start_y = pdf.get_y()
    
    for detail in entry.details:
        if pdf.get_y() > 255:
            pdf.line(31, start_y, 31, pdf.get_y())
            pdf.line(119, start_y, 119, pdf.get_y())
            pdf.line(139, start_y, 139, pdf.get_y())
            pdf.line(159, start_y, 159, pdf.get_y())
            pdf.line(179, start_y, 179, pdf.get_y())
            
            pdf.add_page()
            start_y = pdf.get_y()
            pdf.set_font('helvetica', '', 8)
            
        debit_usd = float(detail.debit) / 6.96 if detail.debit > 0 else 0
        credit_usd = float(detail.credit) / 6.96 if detail.credit > 0 else 0
        
        pdf.cell(20, 5, str(detail.account.code), 0, 0, 'C')
        
        name = detail.account.name
        if len(name) > 60:
            name = name[:57] + "..."
        pdf.cell(88, 5, name.upper(), 0, 0, 'L')
        
        d_str = f"{detail.debit:,.2f}" if detail.debit > 0 else ""
        c_str = f"{detail.credit:,.2f}" if detail.credit > 0 else ""
        pdf.cell(20, 5, d_str, 0, 0, 'R')
        pdf.cell(20, 5, c_str, 0, 0, 'R')
        
        usd_d_str = f"{debit_usd:,.2f}" if debit_usd > 0 else ""
        usd_c_str = f"{credit_usd:,.2f}" if credit_usd > 0 else ""
        pdf.cell(20, 5, usd_d_str, 0, 0, 'R')
        pdf.cell(20, 5, usd_c_str, 0, 1, 'R')
        
        total_debit += detail.debit
        total_credit += detail.credit
        total_debit_usd += debit_usd
        total_credit_usd += credit_usd
        
    last_y = pdf.get_y()
    
    pdf.line(31, start_y, 31, last_y)
    pdf.line(119, start_y, 119, last_y)
    pdf.line(139, start_y, 139, last_y)
    pdf.line(159, start_y, 159, last_y)
    pdf.line(179, start_y, 179, last_y)
    
    pdf.line(11, last_y, 199, last_y)
    
    pdf.set_y(last_y)
    pdf.set_font('helvetica', 'B', 8)
    pdf.cell(108, 6, "SUB TOTAL", 0, 0, 'C')
    pdf.cell(20, 6, f"{total_debit:,.2f}", 1, 0, 'R')
    pdf.cell(20, 6, f"{total_credit:,.2f}", 1, 0, 'R')
    pdf.cell(20, 6, f"{total_debit_usd:,.2f}", 1, 0, 'R')
    pdf.cell(20, 6, f"{total_credit_usd:,.2f}", 1, 1, 'R')
        
    return pdf.output()

class LibroDiarioPDF(FPDF):
    def __init__(self, *args, **kwargs):
        self.start_date_str = kwargs.pop('start_date_str', '')
        self.end_date_str = kwargs.pop('end_date_str', '')
        super().__init__(*args, **kwargs)
        self.set_margins(15, 15, 15)

    def header(self):
        # Header according to template
        # Left side: Company Info
        # It's a bit complex with FPDF to do different alignments on same line,
        # so we use X, Y positioning
        
        # Save current Y
        start_y = self.get_y()
        start_x = self.get_x()
        
        # Company Info (Left)
        self.set_font('helvetica', 'B', 9)
        self.cell(60, 5, 'COMSATELITAL S.R.L.', ln=True, align='L')
        self.set_font('helvetica', '', 9)
        self.cell(60, 5, 'LA PAZ - Bolivia', ln=True, align='L')
        
        # Title (Center)
        # Go back to start Y and draw the title centered over the whole page
        self.set_xy(start_x, start_y)
        self.set_font('helvetica', 'B', 12)
        self.cell(0, 5, 'LIBRO DIARIO', align='C', ln=True)
        self.set_font('helvetica', '', 9)
        self.cell(0, 5, '(Expresado en Bolivianos)', align='C', ln=True)
        
        self.ln(5)
        
        # "BOLIVIANOS" title above Debe and Haber columns
        # Debe is around X=150, Haber is around X=175. So center around 165
        self.set_font('helvetica', 'B', 8)
        self.set_x(140)
        self.cell(40, 5, 'BOLIVIANOS', align='C', ln=True)
        
        # Table Header Row (No borders, just text, with a top and bottom line)
        self.set_font('helvetica', 'B', 8)
        self.line(self.get_x(), self.get_y(), 195, self.get_y()) # Top line
        
        # X: 15 (Fecha), 40 (CODIGO), 75 (DETALLE), 150 (DEBE), 175 (HABER)
        self.cell(20, 5, 'Fecha', 0, 0, 'C')
        self.cell(25, 5, 'CODIGO', 0, 0, 'C')
        self.cell(90, 5, 'DETALLE', 0, 0, 'C')
        self.cell(25, 5, 'DEBE', 0, 0, 'C')
        self.cell(25, 5, 'HABER', 0, 1, 'C')
        
        self.line(self.get_x(), self.get_y(), 195, self.get_y()) # Bottom line
        self.ln(2)

    def footer(self):
        self.set_y(-15)
        self.set_font('helvetica', 'I', 8)
        self.cell(0, 10, f'Página {self.page_no()}/{{nb}}', align='C')


def generate_libro_diario_pdf(entries, start_date_str, end_date_str):
    pdf = LibroDiarioPDF(start_date_str=start_date_str, end_date_str=end_date_str)
    pdf.alias_nb_pages()
    pdf.add_page()
    
    pdf.set_font('helvetica', '', 8)
    
    total_debit = 0
    total_credit = 0
    
    for entry in entries:
        # Check if we need to add a page (to avoid widowed entry headers)
        if pdf.get_y() > 250:
            pdf.add_page()
            
        pdf.set_font('helvetica', 'B', 8)
        type_name = get_voucher_type_name(entry.entry_type)
        entry_desc = f"COMPROBANTE DE {type_name}    Nro.  :{entry.entry_number or entry.id}"
        
        # Row 1: Date | Entry Description
        pdf.cell(20, 5, entry.date.strftime("%d/%m/%Y"), 0, 0, 'C')
        pdf.cell(25, 5, '', 0, 0)
        pdf.cell(90, 5, entry_desc, 0, 1, 'L')
        
        pdf.set_font('helvetica', '', 8)
        
        entry_debit = 0
        entry_credit = 0
        
        for detail in entry.details:
            if pdf.get_y() > 265:
                pdf.add_page()
                
            pdf.cell(20, 5, '', 0, 0)
            pdf.cell(25, 5, str(detail.account.code), 0, 0, 'L')
            
            name = detail.account.name
            if len(name) > 55:
                name = name[:52] + "..."
                
            # If it's a credit, indent a bit (conventional)
            if detail.credit > 0:
                name = "  " + name
                
            pdf.cell(90, 5, name, 0, 0, 'L')
            
            # Print debit or empty
            if detail.debit > 0:
                pdf.cell(25, 5, f"{detail.debit:,.2f}", 0, 0, 'R')
            else:
                pdf.cell(25, 5, "", 0, 0, 'R')
                
            # Print credit or empty
            if detail.credit > 0:
                pdf.cell(25, 5, f"{detail.credit:,.2f}", 0, 1, 'R')
            else:
                pdf.cell(25, 5, "", 0, 1, 'R')
                
            entry_debit += detail.debit
            entry_credit += detail.credit
            
        # Print entry total (Glosa + Totals row)
        if pdf.get_y() > 265:
            pdf.add_page()
            
        pdf.set_font('helvetica', '', 8)
        pdf.cell(20, 5, '', 0, 0)
        pdf.cell(25, 5, '', 0, 0)
        
        desc = entry.description or ''
        # Multi_cell handles word wrap, but to align the amounts on the same line
        # as the LAST line of the multi_cell, or just the first line, we need some trickery.
        # Let's truncate desc for simplicity or use a fixed height
        if len(desc) > 65:
             desc = desc[:62] + "..."
             
        pdf.cell(90, 5, desc, 0, 0, 'L')
        
        pdf.cell(25, 5, f"{entry_debit:,.2f}", 0, 0, 'R')
        pdf.cell(25, 5, f"{entry_credit:,.2f}", 0, 1, 'R')
        
        total_debit += entry_debit
        total_credit += entry_credit
        # Add a tiny space between entries
        pdf.ln(2)
        
    # Grand Total
    if pdf.get_y() > 265:
        pdf.add_page()
        
    pdf.ln(5)
    pdf.set_font('helvetica', 'B', 8)
    pdf.cell(45, 5, '', 0, 0)
    pdf.cell(90, 5, "TOTAL GENERAL DEL DIARIO", 0, 0, 'R')
    pdf.cell(25, 5, f"{total_debit:,.2f}", 0, 0, 'R')
    pdf.cell(25, 5, f"{total_credit:,.2f}", 0, 1, 'R')
        
    return pdf.output()

def generate_libro_diario_excel(entries, start_date_str, end_date_str):
    wb = Workbook()
    ws = wb.active
    ws.title = "Libro Diario"
    
    # Styling
    bold_font = Font(bold=True, size=10)
    center_align = Alignment(horizontal='center')
    border = Border(bottom=Side(style='thin'))
    
    # Headers
    ws['A1'] = 'COMSATELITAL S.R.L.'
    ws['A1'].font = bold_font
    ws['A2'] = 'LA PAZ - BOLIVIA'
    
    ws.merge_cells('A4:H4')
    ws['A4'] = 'LIBRO DIARIO'
    ws['A4'].font = Font(bold=True, size=14)
    ws['A4'].alignment = center_align
    
    ws.merge_cells('A5:H5')
    ws['A5'] = f'Del {start_date_str} al {end_date_str}'
    ws['A5'].alignment = center_align
    
    ws['A6'] = '(Expresado en Bolivianos)'
    ws['A6'].alignment = center_align
    ws.merge_cells('A6:H6')
    
    # Currency Header
    ws['G7'] = 'BOLIVIANOS'
    ws['G7'].font = bold_font
    ws['G7'].alignment = center_align
    ws.merge_cells('G7:H7')
    
    # Table Header
    headers = ['Nro', 'Tipo', 'Fecha', 'Código', 'Detalle', '', 'Debe', 'Haber']
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=8, column=col_num, value=header)
        cell.font = bold_font
        cell.alignment = center_align
        cell.border = border
        
    current_row = 9
    total_debit = 0
    total_credit = 0
    nro_counter = 1
    
    for entry in entries:
        # Entry header block
        ws.cell(row=current_row, column=1, value=nro_counter)
        ws.cell(row=current_row, column=2, value=entry.entry_type)
        ws.cell(row=current_row, column=3, value=entry.date.strftime("%d/%m/%Y"))
        
        type_name = get_voucher_type_name(entry.entry_type)
        entry_title = f"COMPROBANTE DE {type_name} Nro. :{entry.entry_number or entry.id}"
        ws.cell(row=current_row, column=5, value=clean_string(entry_title)).font = bold_font
        
        current_row += 1
        entry_debit = 0
        entry_credit = 0
        
        for detail in entry.details:
            ws.cell(row=current_row, column=4, value=clean_string(str(detail.account.code)))
            ws.cell(row=current_row, column=5, value=clean_string(detail.account.name))
            
            if detail.debit > 0:
                ws.cell(row=current_row, column=7, value=float(detail.debit))
                entry_debit += detail.debit
            if detail.credit > 0:
                ws.cell(row=current_row, column=8, value=float(detail.credit))
                entry_credit += detail.credit
            
            current_row += 1
            
        # Entry Footer (Glosa and Subtotals)
        ws.cell(row=current_row, column=5, value=clean_string(entry.description or '')).font = Font(italic=True)
        ws.cell(row=current_row, column=7, value=float(entry_debit)).font = bold_font
        ws.cell(row=current_row, column=8, value=float(entry_credit)).font = bold_font
        
        total_debit += entry_debit
        total_credit += entry_credit
        nro_counter += 1
        current_row += 2 # Space between entries
        
    # Final Total
    ws.cell(row=current_row, column=5, value="TOTAL GENERAL DEL DIARIO").font = bold_font
    ws.cell(row=current_row, column=5).alignment = Alignment(horizontal='right')
    ws.cell(row=current_row, column=7, value=float(total_debit)).font = bold_font
    ws.cell(row=current_row, column=8, value=float(total_credit)).font = bold_font
    
    # Adjust column widths
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 50
    ws.column_dimensions['G'].width = 15
    ws.column_dimensions['H'].width = 15
    
    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()

def generate_tax_audit_excel(report):
    wb = Workbook()
    ws = wb.active
    ws.title = "Auditoria Tributaria"
    
    # Styles
    title_font = Font(bold=True, size=14, color="000080")
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="000080", end_color="000080", fill_type="solid")
    border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    
    # Header Info
    ws['A1'] = "AUDITORIA TRIBUTARIA"
    ws['A1'].font = title_font
    ws['A2'] = f"EMPRESA: {report.taxpayer_name}"
    ws['A3'] = f"NIT: {report.taxpayer_nit}"
    ws['A4'] = "Vaciado F-200 V.6 - Impuesto al Valor Agregado"
    ws['A4'].font = Font(bold=True, color="FF0000")
    ws['A5'] = f"Revisión Tributaria: {report.year}"
    
    # Table Header
    months = ["Enero", "Febrero", "Marzo", "Abril", "Mayo", "Junio", "Julio", "Agosto", "Septiembre", "Octubre", "Noviembre", "Diciembre"]
    cols = ["DETALLE", "COD"] + months + ["TOTAL"]
    
    current_row = 7
    for col_idx, col_name in enumerate(cols, 1):
        cell = ws.cell(row=current_row, column=col_idx, value=col_name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = border

    current_row += 1

    # Meta Rows (Date, Order Number)
    # Fecha de presentación
    ws.cell(row=current_row, column=1, value="Fecha de presentación").border = border
    ws.cell(row=current_row, column=2, value="").border = border
    for m in range(12):
        d = report.presentation_dates[m]
        val = d.strftime("%d/%m/%Y") if d else "-"
        ws.cell(row=current_row, column=3+m, value=val).border = border
    ws.cell(row=current_row, column=15, value="").border = border
    current_row += 1

    # Número de orden
    ws.cell(row=current_row, column=1, value="Número de orden").border = border
    ws.cell(row=current_row, column=2, value="").border = border
    for m in range(12):
        val = report.transaction_numbers[m] or "-"
        ws.cell(row=current_row, column=3+m, value=val).border = border
    ws.cell(row=current_row, column=15, value="").border = border
    current_row += 1

    # Data Rows
    last_rubric = ""
    for row in report.rows:
        if row.rubric != last_rubric:
            ws.cell(row=current_row, column=1, value=row.rubric).font = Font(bold=True)
            ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=15)
            last_rubric = row.rubric
            current_row += 1
        
        ws.cell(row=current_row, column=1, value=row.label).border = border
        ws.cell(row=current_row, column=2, value=row.field_code).border = border
        for m_idx, val in enumerate(row.months):
            c = ws.cell(row=current_row, column=3+m_idx, value=float(val))
            c.border = border
            c.number_format = '#,##0.00'
        
        tot_c = ws.cell(row=current_row, column=15, value=float(row.total))
        tot_c.border = border
        tot_c.font = Font(bold=True)
        tot_c.number_format = '#,##0.00'
        current_row += 1

    # Adjust widths
    ws.column_dimensions['A'].width = 60
    ws.column_dimensions['B'].width = 8
    for col_l in ['C','D','E','F','G','H','I','J','K','L','M','N','O']:
        ws.column_dimensions[col_l].width = 12

    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()

class TaxAuditPDF(FPDF):
    def __init__(self, *args, **kwargs):
        self.report = kwargs.pop('report', None)
        super().__init__(*args, **kwargs)
        self.set_margins(10, 10, 10)

    def header(self):
        if not self.report: return
        self.set_font('helvetica', 'B', 12)
        self.cell(0, 6, "AUDITORIA TRIBUTARIA", ln=True)
        self.set_font('helvetica', '', 9)
        self.cell(0, 5, f"EMPRESA: {self.report.taxpayer_name}", ln=True)
        self.cell(0, 5, f"NIT: {self.report.taxpayer_nit}", ln=True)
        self.set_text_color(200, 0, 0)
        self.set_font('helvetica', 'B', 10)
        self.cell(0, 5, "Vaciado F-200 V.6 - Impuesto al Valor Agregado", ln=True)
        self.set_text_color(0, 0, 0)
        self.set_font('helvetica', '', 9)
        self.cell(0, 5, f"Revisión Tributaria: {self.report.year}", ln=True)
        self.ln(5)
        
        # Table Header
        self.set_font('helvetica', 'B', 6)
        self.set_fill_color(0, 0, 128)
        self.set_text_color(255, 255, 255)
        
        # A4 Landscape width is ~297mm. With 10mm margins = 277mm usable.
        self.cell(70, 5, "DETALLE", 1, 0, 'C', True)
        self.cell(10, 5, "COD", 1, 0, 'C', True)
        months = ["Ene", "Feb", "Mar", "Abr", "May", "Jun", "Jul", "Ago", "Sep", "Oct", "Nov", "Dic"]
        for m in months:
            self.cell(15, 5, m, 1, 0, 'C', True)
        self.cell(17, 5, "TOTAL", 1, 1, 'C', True)
        self.set_text_color(0, 0, 0)

    def footer(self):
        self.set_y(-15)
        self.set_font('helvetica', 'I', 6)
        self.cell(0, 10, f'Página {self.page_no()}/{{nb}}', align='C')

def generate_tax_audit_pdf(report):
    pdf = TaxAuditPDF(orientation='L', unit='mm', format='A4', report=report)
    pdf.alias_nb_pages()
    pdf.add_page()
    
    pdf.set_font('helvetica', '', 6)
    
    # Meta rows
    pdf.set_fill_color(245, 245, 245)
    # Fecha Pres
    pdf.cell(70, 4, "Fecha de presentación", 1, 0, 'L', True)
    pdf.cell(10, 4, "", 1, 0, 'C', True)
    for d in report.presentation_dates:
        val = d.strftime("%d/%m/%y") if d else "-"
        pdf.cell(15, 4, val, 1, 0, 'C')
    pdf.cell(17, 4, "", 1, 1)

    # Nro Orden
    pdf.cell(70, 4, "Número de orden", 1, 0, 'L', True)
    pdf.cell(10, 4, "", 1, 0, 'C', True)
    for n in report.transaction_numbers:
        val = (n[-8:] if n and len(n) > 8 else n) or "-"
        pdf.cell(15, 4, val, 1, 0, 'C')
    pdf.cell(17, 4, "", 1, 1)

    last_rubric = ""
    for row in report.rows:
        if row.rubric != last_rubric:
            pdf.set_font('helvetica', 'B', 7)
            pdf.set_fill_color(230, 230, 250)
            pdf.cell(277, 5, row.rubric, 1, 1, 'L', True)
            pdf.set_font('helvetica', '', 6)
            last_rubric = row.rubric
        
        label = row.label
        if len(label) > 65: label = label[:62] + "..."
        
        pdf.cell(70, 4, label, 1)
        pdf.cell(10, 4, row.field_code, 1, 0, 'C')
        for val in row.months:
            pdf.cell(15, 4, f"{val:,.0f}", 1, 0, 'R')
        
        pdf.set_font('helvetica', 'B', 6)
        pdf.cell(17, 4, f"{row.total:,.0f}", 1, 1, 'R')
        pdf.set_font('helvetica', '', 6)

    return pdf.output()

def generate_resumen_formulario_200_excel(report):
    import openpyxl
    import os
    
    template_path = r'c:\sicjac\reportes\ResumenFormulario200.xlsx'
    if not os.path.exists(template_path):
        # Fallback to creating a new workbook if template is missing
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Resumen F200"
    else:
        wb = openpyxl.load_workbook(template_path)
        ws = wb.active

    # Row 3 contains the field codes (C13, C14, etc.)
    # We need to map these to the report data
    field_to_col = {}
    for col in range(2, ws.max_column + 1):
        val = ws.cell(row=3, column=col).value
        if val and str(val).startswith('C'):
            code = str(val).replace('C', '').strip()
            field_to_col[code] = col
    
    # Summary columns mapping by numeric code
    code_to_summary_col = {
        "909": 52,
        "693": 53,
        "635": 54,
        "648": 55,
        "1001": 56,
        "592": 57
    }
    
    # Months rows: 4 to 15 (Enero to Diciembre)
    for m_idx in range(12):
        row_num = 4 + m_idx
        
        # Populate all field codes and mapped summary labels
        for row in report.rows:
            code = row.field_code.replace('C', '').replace('Casilla ', '').strip()
            
            # Map to its primary column from row 3
            if code in field_to_col:
                col_num = field_to_col[code]
                ws.cell(row=row_num, column=col_num, value=float(row.months[m_idx]))
            
            # Also populate labels at the end of the row (Cols 52-57)
            if code in code_to_summary_col:
                col_num = code_to_summary_col[code]
                ws.cell(row=row_num, column=col_num, value=float(row.months[m_idx]))

    # Add TOTAL row (Row 16)
    row_total = 16
    ws.cell(row=row_total, column=1, value="TOTAL")
    from openpyxl.styles import Font
    ws.cell(row=row_total, column=1).font = Font(bold=True)
    
    # Calculate totals for each column that was populated
    for code, col_num in field_to_col.items():
        total_val = 0
        for row in report.rows:
            clean_row_code = str(row.field_code).replace('C', '').replace('Casilla ', '').strip()
            if clean_row_code == code:
                try:
                    total_val = sum(float(m) if m is not None else 0 for m in row.months)
                except (ValueError, TypeError):
                    total_val = 0
                break
        ws.cell(row=row_total, column=col_num, value=total_val)
        ws.cell(row=row_total, column=col_num).font = Font(bold=True)
        
    for code, col_num in code_to_summary_col.items():
        total_val = 0
        for row in report.rows:
            clean_row_code = str(row.field_code).replace('C', '').replace('Casilla ', '').strip()
            if clean_row_code == code:
                try:
                    total_val = sum(float(m) if m is not None else 0 for m in row.months)
                except (ValueError, TypeError):
                    total_val = 0
                break
        ws.cell(row=row_total, column=col_num, value=total_val)
        ws.cell(row=row_total, column=col_num).font = Font(bold=True)

    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()
